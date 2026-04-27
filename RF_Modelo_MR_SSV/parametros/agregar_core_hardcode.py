"""Actualiza parametros_mr_ssv.xlsx:

- Agrega/actualiza la hoja ``CORE_HARDCODE`` con los 4 valores de CORE
  hardcodeados por metodologia (leidos de la hoja ``Montos`` del XLSM
  heredado ``MT_SSV.XLSM`` -- celdas D3:D6).
- Preserva las hojas existentes (``CUOTAS_SSV``, ``DISTR_CORE_SSV_R13``,
  ``FACTORES``).
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

AQUI = Path(__file__).resolve().parent
RUTA_PARAM = AQUI / "parametros_mr_ssv.xlsx"

# Valores observados en MT_SSV.XLSM hoja Montos (D3:D6) al 2026-04-17.
# Son definidos por metodologia (cambian trimestralmente segun cierre).
CORE_HARDCODE = pd.DataFrame([
    {"COD_SUB_PRO_MODELO": "CTA_CTE", "MONTO_CORE_GESTION_MO": 1_297_659_000_000.0,
     "MONEDA": "CLP", "FUENTE": "MT_SSV.XLSM Montos!D3 al 2026-04-17"},
    {"COD_SUB_PRO_MODELO": "CTA_VTA", "MONTO_CORE_GESTION_MO":   125_122_000_000.0,
     "MONEDA": "CLP", "FUENTE": "MT_SSV.XLSM Montos!D4 al 2026-04-17"},
    {"COD_SUB_PRO_MODELO": "AGD",     "MONTO_CORE_GESTION_MO":         5_006_933.0,
     "MONEDA": "CLF", "FUENTE": "MT_SSV.XLSM Montos!D5 al 2026-04-17"},
    {"COD_SUB_PRO_MODELO": "AGI",     "MONTO_CORE_GESTION_MO":         1_065_518.0,
     "MONEDA": "CLF", "FUENTE": "MT_SSV.XLSM Montos!D6 al 2026-04-17"},
])


def main() -> None:
    if not RUTA_PARAM.exists():
        raise FileNotFoundError(RUTA_PARAM)

    wb = load_workbook(RUTA_PARAM)
    hojas_antes = wb.sheetnames
    print(f"Hojas antes: {hojas_antes}")
    if "CORE_HARDCODE" in wb.sheetnames:
        del wb["CORE_HARDCODE"]
    wb.save(RUTA_PARAM)

    # Usamos openpyxl engine con mode='a' para conservar el resto.
    with pd.ExcelWriter(RUTA_PARAM, engine="openpyxl", mode="a") as w:
        CORE_HARDCODE.to_excel(w, sheet_name="CORE_HARDCODE", index=False)

    # Verificacion
    wb2 = load_workbook(RUTA_PARAM)
    print(f"Hojas despues: {wb2.sheetnames}")
    print(CORE_HARDCODE.to_string(index=False))


if __name__ == "__main__":
    main()
